import random
import re
import openpyxl as op

from tables.Event import Event


def excel_main(hotels, period, event_objects, start_id):
    events = read_text_file()
    create_events_table(events, hotels, period, event_objects, start_id + 1)
    add_events_to_excel_file(event_objects, start_id + 1)


def read_text_file():
    category = ''
    events = {}

    with open('static/wydarzenia.txt', encoding='utf8') as file:
        content = file.readlines()
        for line in content:
            if re.search('KATEGORIA', line):
                line = line.replace('KATEGORIA - ', '')
                line = line.replace('\n', '')
                category = line
                events[category] = []

            elif re.search('NAZWA_OPIS', line):
                line = line.replace('NAZWA_OPIS - ', '')
                line = line.replace('\n', '')
                splitted_text = line.split(' - ')
                events[category].append((splitted_text[0], splitted_text[1]))

    return events


def create_events_table(events, hotels, period, event_objects, start_id):
    events_list = []
    dates = []

    for category in events:
        for event in events[category]:
            events_list.append((event[0], event[1], category))

    for _ in range(1000):
        event_number = random.randrange(len(events_list))
        event_object = events_list[event_number]
        event = Event(len(event_objects) + start_id, hotels, period, event_object[0], event_object[1], event_object[2], dates)

        event_objects.append(event)


def add_events_to_excel_file(event_objects, start_id):
    wb = op.load_workbook('static/events_and_room_types.xlsx')
    ws = wb['Wydarzenie']
    i = start_id + 1

    for event in event_objects:
        ws['A' + str(i)] = event.id
        ws['B' + str(i)] = int(event.hotel_id)
        ws['C' + str(i)] = event.date
        ws['D' + str(i)] = event.name
        ws['E' + str(i)] = event.description
        ws['F' + str(i)] = event.type
        ws['G' + str(i)] = event.max_no_of_persons
        ws['H' + str(i)] = event.no_of_participants

        i += 1

    wb.save('static/events_and_room_types.xlsx')
